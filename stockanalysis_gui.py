"""
stockanalysis_gui.py
────────────────────
GUI for the stockanalysis.com scraper.
Uses only tkinter (built into Python — no extra install needed).

SETUP
-----
1. Make sure dependencies are installed:
       pip install requests pandas beautifulsoup4 lxml

2. Run this file:
       python stockanalysis_gui.py
"""

import os
import time
import threading
import tkinter as tk
from datetime import datetime
from tkinter import ttk, filedialog, messagebox

import requests
import pandas as pd
from bs4 import BeautifulSoup


# ── Scraper logic ────────────────────────────────────────────────────────────

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
}

STATEMENTS = [
    "overview",
    "income-statement",
    "balance-sheet",
    "cash-flow-statement",
    "ratios",
]

DELAY_SECONDS = 1.5


def build_url(ticker: str, statement: str, period: str) -> str:
    base = f"https://stockanalysis.com/stocks/{ticker.lower()}/financials"
    path_map = {
        "overview":            f"https://stockanalysis.com/stocks/{ticker.lower()}/",
        "income-statement":    base,
        "balance-sheet":       f"{base}/balance-sheet/",
        "cash-flow-statement": f"{base}/cash-flow-statement/",
        "ratios":              f"{base}/ratios/",
    }
    url = path_map[statement]
    if period == "quarterly" and statement != "overview":
        url += "?p=quarterly"
    return url


def clean_value(td) -> str:
    """Strip inline % change spans from a table cell, keeping only the primary value."""
    import copy
    td = copy.copy(td)
    for span in td.find_all("span"):
        cls  = " ".join(span.get("class") or [])
        text = span.get_text(strip=True)
        is_pct_class = any(k in cls for k in ["chg", "change", "percent", "pos", "neg", "up", "down", "green", "red"])
        is_pct_text  = bool(text) and text[-1] == "%" and text[0] in ("+", "-")
        if is_pct_class or is_pct_text:
            span.decompose()
    return td.get_text(strip=True)


def scrape_overview(soup: BeautifulSoup) -> pd.DataFrame | None:
    rows = []
    for table in soup.find_all("table"):
        for tr in table.find_all("tr"):
            tds = tr.find_all(["td", "th"])
            if len(tds) >= 2:
                label = tds[0].get_text(strip=True)
                value = clean_value(tds[1])
                if label and value:
                    rows.append([label, value])
    seen, unique = set(), []
    for r in rows:
        if r[0] and r[0] not in seen:
            seen.add(r[0])
            unique.append(r)
    return pd.DataFrame(unique, columns=["Metric", "Value"]) if unique else None


# Sentinel returned when a network/HTTP error occurs (bad ticker, rate-limited, etc.)
# Distinct from None which means "page loaded but no table found"
class ScrapeError:
    def __init__(self, reason: str):
        self.reason = reason


def scrape_table(ticker: str, statement: str, period: str):
    if statement == "overview" and period == "quarterly":
        return None
    url = build_url(ticker, statement, period)
    try:
        resp = requests.get(url, headers=REQUEST_HEADERS, timeout=20)
        resp.raise_for_status()
    except requests.exceptions.HTTPError as e:
        return ScrapeError(f"HTTP {e.response.status_code} — ticker may be invalid")
    except requests.RequestException as e:
        return ScrapeError(f"Network error: {e}")
    soup = BeautifulSoup(resp.text, "lxml")
    if statement == "overview":
        return scrape_overview(soup)
    table = (
        soup.find("table", {"class": lambda c: c and "financial-table" in c})
        or soup.find("table", {"class": lambda c: c and "w-full" in c})
        or soup.find("table")
    )
    if not table:
        return None
    # Use ONLY the first <tr> in thead — the second row is "Period Ending" dates
    thead = table.find("thead")
    if not thead:
        return None
    first_header_row = thead.find("tr")
    if not first_header_row:
        return None
    headers = [th.get_text(strip=True) for th in first_header_row.find_all("th")]
    if not headers:
        return None
    rows = []
    tbody = table.find("tbody")
    if tbody:
        for tr in tbody.find_all("tr"):
            cells = [td.get_text(strip=True) for td in tr.find_all(["td", "th"])]
            if not cells or cells[0] == "Period Ending" or not any(c for c in cells):
                continue
            rows.append(cells)
    if not rows:
        return None
    # Pin columns to header length to avoid trailing empty columns
    max_cols = len(headers)
    rows = [r[:max_cols] + [""] * max(0, max_cols - len(r)) for r in rows]
    return pd.DataFrame(rows, columns=headers)


def _write_sheet(ws, df: pd.DataFrame, ticker: str, statement: str):
    """
    Write a DataFrame to an openpyxl worksheet with:
    - Metadata rows: Ticker, Date, Time, Currency
    - A blank separator row
    - The data table with auto-fitted column widths
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    HEADER_FILL  = PatternFill("solid", fgColor="1A1D27")
    HEADER_FONT  = Font(bold=True, color="E8EAF0", size=10)
    META_FONT    = Font(color="7B7F93", size=9, italic=True)
    LABEL_FONT   = Font(bold=True, color="7B7F93", size=9)

    now = dt.now()
    meta = [
        ("Ticker",   ticker.upper()),
        ("Date",     now.strftime("%B %d, %Y")),
        ("Time",     now.strftime("%I:%M %p")),
        ("Currency", "USD"),
        ("Source",   "stockanalysis.com"),
    ]

    # Write metadata rows
    for label, value in meta:
        ws.append([label, value])
        row = ws.max_row
        ws.cell(row, 1).font = LABEL_FONT
        ws.cell(row, 2).font = META_FONT

    ws.append([])  # blank separator

    # Write column headers
    header_row = ws.max_row + 1
    ws.append(list(df.columns))
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws.cell(header_row, col_idx)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for _, row_data in df.iterrows():
        ws.append(list(row_data))

    # Auto-fit column widths based on all content
    for col_idx, col_cells in enumerate(ws.columns, 1):
        max_len = 0
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)


def save_file(df: pd.DataFrame, output_dir: str, ticker: str, statement: str, period: str, fmt: str) -> str:
    import openpyxl
    os.makedirs(output_dir, exist_ok=True)
    ext = "xlsx" if fmt == "xlsx" else "csv"
    base = f"{ticker.upper()}_overview" if statement == "overview" \
        else f"{ticker.upper()}_{statement}_{period}"
    filepath = os.path.join(output_dir, f"{base}.{ext}")
    if fmt == "xlsx":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = statement[:31]
        _write_sheet(ws, df, ticker, statement)
        wb.save(filepath)
    else:
        df.to_csv(filepath, index=False)
    return filepath


def save_combined(sheets: dict, output_dir: str, ticker: str) -> str:
    """Save all statements as tabs in a single Excel workbook with metadata and auto-width."""
    import openpyxl
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, f"{ticker.upper()}_combined.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet
    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        _write_sheet(ws, df, ticker, sheet_name)
    wb.save(filepath)
    return filepath


# ── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".stockscraper_config.json")

    def __init__(self):
        super().__init__()
        self.title("StockAnalysis Scraper")
        self.resizable(False, False)
        self.configure(bg="#0f1117")
        self._build_ui()
        self._load_config()
        self.bind("<Return>", lambda event: self._start_scrape())
        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _load_config(self):
        import json
        try:
            with open(self.CONFIG_PATH) as f:
                cfg = json.load(f)
            if "output_dir" in cfg:
                self.folder_var.set(cfg["output_dir"])
        except Exception:
            pass

    def _save_config(self):
        import json
        try:
            with open(self.CONFIG_PATH, "w") as f:
                json.dump({"output_dir": self.folder_var.get().strip()}, f)
        except Exception:
            pass

    def _on_close(self):
        self._save_config()
        self.destroy()

    def _build_ui(self):
        DARK   = "#0f1117"
        PANEL  = "#1a1d27"
        BORDER = "#2a2d3a"
        ACCENT = "#4f8ef7"
        TEXT   = "#e8eaf0"
        MUTED  = "#7b7f93"
        GREEN  = "#3ecf8e"
        RED    = "#f76f6f"

        self._colors = dict(DARK=DARK, PANEL=PANEL, BORDER=BORDER,
                            ACCENT=ACCENT, TEXT=TEXT, MUTED=MUTED,
                            GREEN=GREEN, RED=RED)

        pad = dict(padx=20, pady=10)

        # ── Header ──
        hdr = tk.Frame(self, bg=DARK)
        hdr.pack(fill="x", padx=24, pady=(24, 4))
        tk.Label(hdr, text="StockAnalysis.com Scraper",
                 font=("Helvetica", 18, "bold"), fg=TEXT, bg=DARK).pack(side="left")

        user_frame = tk.Frame(hdr, bg=DARK)
        user_frame.pack(side="right")
        tk.Label(user_frame, text="Nick Barran",
                 font=("Helvetica", 10, "bold"), fg=TEXT, bg=DARK).pack(anchor="e")
        tk.Label(user_frame, text="nbarran@uw.edu",
                 font=("Helvetica", 9), fg=MUTED, bg=DARK).pack(anchor="e")

        # ── Main card ──
        card = tk.Frame(self, bg=PANEL, bd=0, highlightthickness=1,
                        highlightbackground=BORDER)
        card.pack(padx=24, pady=8, fill="both")

        # Tickers
        tk.Label(card, text="Tickers", font=("Helvetica", 11, "bold"),
                 fg=TEXT, bg=PANEL).grid(row=0, column=0, sticky="w", padx=20, pady=(18, 4))
        tk.Label(card, text="Separate with commas — e.g. AAPL, MSFT, NVDA",
                 font=("Helvetica", 9), fg=MUTED, bg=PANEL).grid(
                 row=1, column=0, sticky="w", padx=20, pady=(0, 4))

        self.ticker_var = tk.StringVar()
        ticker_entry = tk.Entry(card, textvariable=self.ticker_var,
                                font=("Helvetica", 12), bg="#252836",
                                fg=TEXT, insertbackground=TEXT,
                                relief="flat", bd=8, width=38)
        ticker_entry.grid(row=2, column=0, columnspan=2, padx=20, pady=(0, 14), sticky="ew")

        # Period
        tk.Label(card, text="Period", font=("Helvetica", 11, "bold"),
                 fg=TEXT, bg=PANEL).grid(row=3, column=0, sticky="w", padx=20, pady=(4, 8))

        period_frame = tk.Frame(card, bg=PANEL)
        period_frame.grid(row=4, column=0, columnspan=2, sticky="w", padx=20, pady=(0, 14))

        self.period_var = tk.StringVar(value="annual")
        for val, label in [("annual", "Annual"), ("quarterly", "Quarterly"), ("both", "Both")]:
            rb = tk.Radiobutton(period_frame, text=label, variable=self.period_var,
                                value=val, font=("Helvetica", 11),
                                fg=TEXT, bg=PANEL, selectcolor=PANEL,
                                activebackground=PANEL, activeforeground=ACCENT,
                                indicatoron=0, relief="flat", bd=0,
                                padx=14, pady=6, cursor="hand2")
            rb.pack(side="left", padx=(0, 8))
        self._style_radiobuttons(period_frame, ACCENT, BORDER, TEXT)

        # Output Format
        tk.Label(card, text="Format", font=("Helvetica", 11, "bold"),
                 fg=TEXT, bg=PANEL).grid(row=5, column=0, sticky="w", padx=20, pady=(4, 8))

        format_frame = tk.Frame(card, bg=PANEL)
        format_frame.grid(row=6, column=0, columnspan=2, sticky="w", padx=20, pady=(0, 14))

        self.format_var = tk.StringVar(value="xlsx")
        for val, lbl in [("csv", "CSV"), ("xlsx", "Excel (.xlsx)")]:
            rb = tk.Radiobutton(format_frame, text=lbl, variable=self.format_var,
                                value=val, font=("Helvetica", 11),
                                fg=TEXT, bg=PANEL, selectcolor=PANEL,
                                activebackground=PANEL, activeforeground=ACCENT,
                                indicatoron=0, relief="flat", bd=0,
                                padx=14, pady=6, cursor="hand2")
            rb.pack(side="left", padx=(0, 8))
        self._style_radiobuttons(format_frame, ACCENT, BORDER, TEXT)

        # Output Mode
        tk.Label(card, text="Output Mode", font=("Helvetica", 11, "bold"),
                 fg=TEXT, bg=PANEL).grid(row=7, column=0, sticky="w", padx=20, pady=(4, 8))

        mode_frame = tk.Frame(card, bg=PANEL)
        mode_frame.grid(row=8, column=0, columnspan=2, sticky="w", padx=20, pady=(0, 14))

        self.mode_var = tk.StringVar(value="combined")
        for val, lbl in [("combined", "Combined Excel (one file per ticker)"),
                         ("separate", "Separate files (one per statement)")]:
            rb = tk.Radiobutton(mode_frame, text=lbl, variable=self.mode_var,
                                value=val, font=("Helvetica", 11),
                                fg=TEXT, bg=PANEL, selectcolor=PANEL,
                                activebackground=PANEL, activeforeground=ACCENT,
                                indicatoron=0, relief="flat", bd=0,
                                padx=14, pady=6, cursor="hand2")
            rb.pack(side="left", padx=(0, 8))
        self._style_radiobuttons_var(mode_frame, self.mode_var, ACCENT, BORDER, TEXT)

        # Statements
        tk.Label(card, text="Statements", font=("Helvetica", 11, "bold"),
                 fg=TEXT, bg=PANEL).grid(row=9, column=0, sticky="w", padx=20, pady=(4, 8))

        stmt_frame = tk.Frame(card, bg=PANEL)
        stmt_frame.grid(row=10, column=0, columnspan=2, sticky="w", padx=20, pady=(0, 14))

        self.stmt_vars = {}
        stmt_options = [
            ("overview",            "Overview"),
            ("income-statement",    "Income Statement"),
            ("balance-sheet",       "Balance Sheet"),
            ("cash-flow-statement", "Cash Flow Statement"),
            ("ratios",              "Ratios"),
        ]
        for i, (val, lbl) in enumerate(stmt_options):
            var = tk.BooleanVar(value=True)
            self.stmt_vars[val] = var
            cb = tk.Checkbutton(stmt_frame, text=lbl, variable=var,
                                font=("Helvetica", 11), fg=TEXT, bg=PANEL,
                                selectcolor=ACCENT, activebackground=PANEL,
                                activeforeground=TEXT, relief="flat", bd=0,
                                cursor="hand2")
            row, col = divmod(i, 3)
            cb.grid(row=row, column=col, sticky="w", padx=(0, 20), pady=2)

        # Output folder
        tk.Label(card, text="Output Folder", font=("Helvetica", 11, "bold"),
                 fg=TEXT, bg=PANEL).grid(row=11, column=0, sticky="w", padx=20, pady=(4, 4))

        folder_frame = tk.Frame(card, bg=PANEL)
        folder_frame.grid(row=12, column=0, columnspan=2, sticky="ew", padx=20, pady=(0, 18))

        self.folder_var = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Desktop", "stockscraper"))
        folder_entry = tk.Entry(folder_frame, textvariable=self.folder_var,
                                font=("Helvetica", 10), bg="#252836",
                                fg=TEXT, insertbackground=TEXT,
                                relief="flat", bd=6, width=30)
        folder_entry.pack(side="left", fill="x", expand=True)

        browse_btn = tk.Button(folder_frame, text="Browse", font=("Helvetica", 10),
                               bg=BORDER, fg=TEXT, relief="flat", bd=0,
                               padx=10, pady=5, cursor="hand2",
                               command=self._browse_folder)
        browse_btn.pack(side="left", padx=(8, 0))

        tk.Frame(card, bg=BORDER, height=1).grid(row=13, column=0, columnspan=2,
                                                  sticky="ew", padx=0, pady=0)

        # Run button
        self.run_btn = tk.Button(card, text="▶  Run Scraper",
                                 font=("Helvetica", 13, "bold"),
                                 bg=ACCENT, fg="white", relief="flat", bd=0,
                                 padx=0, pady=14, cursor="hand2",
                                 command=self._start_scrape)
        self.run_btn.grid(row=14, column=0, columnspan=2, sticky="ew",
                          padx=20, pady=16)

        # ── Collapsible Log ──
        self.log_visible = tk.BooleanVar(value=False)

        log_header = tk.Frame(self, bg=PANEL, bd=0, highlightthickness=1,
                              highlightbackground=BORDER)
        log_header.pack(padx=24, pady=(0, 0), fill="x")

        def toggle_log():
            if self.log_visible.get():
                log_body.pack_forget()
                toggle_btn.config(text="▶  Log")
                self.log_visible.set(False)
            else:
                log_body.pack(padx=24, pady=(0, 4), fill="both", expand=True)
                toggle_btn.config(text="▼  Log")
                self.log_visible.set(True)

        toggle_btn = tk.Button(log_header, text="▶  Log",
                               font=("Helvetica", 10, "bold"),
                               fg=MUTED, bg=PANEL, relief="flat", bd=0,
                               padx=16, pady=8, cursor="hand2",
                               activebackground=PANEL, activeforeground=TEXT,
                               command=toggle_log)
        toggle_btn.pack(side="left")

        log_body = tk.Frame(self, bg=PANEL, bd=0, highlightthickness=1,
                            highlightbackground=BORDER)
        # Not packed initially — shown on toggle

        self.log_text = tk.Text(log_body, height=12, bg="#0d0f18", fg=TEXT,
                                font=("Courier", 10), relief="flat", bd=0,
                                state="disabled", wrap="word",
                                padx=12, pady=8)
        self.log_text.pack(fill="both", expand=True, padx=1, pady=1)

        self.log_text.tag_config("ok",    foreground=GREEN)
        self.log_text.tag_config("fail",  foreground=RED)
        self.log_text.tag_config("info",  foreground=ACCENT)
        self.log_text.tag_config("muted", foreground=MUTED)

        # Auto-expand log when scraping starts
        self._toggle_log_fn = toggle_log
        self._log_visible_var = self.log_visible

        # Progress bar — three colour states: default (blue), done (green), error (red)
        style = ttk.Style(self)
        style.theme_use("default")
        for name, color in [
            ("blue.Horizontal.TProgressbar",  ACCENT),
            ("green.Horizontal.TProgressbar", GREEN),
            ("red.Horizontal.TProgressbar",   RED),
        ]:
            style.configure(name, troughcolor=BORDER, background=color, thickness=4)
        self.progress = ttk.Progressbar(self, style="blue.Horizontal.TProgressbar",
                                        mode="determinate")
        self.progress.pack(fill="x", padx=24, pady=(4, 24))

    def _set_progress_color(self, color: str):
        """Switch progress bar colour: 'blue', 'green', or 'red'."""
        self.progress.config(style=f"{color}.Horizontal.TProgressbar")

    def _style_radiobuttons(self, frame, accent, border, text):
        """Apply toggle-button styling to radio buttons (tracks period_var)."""
        def update_styles(event=None):
            for rb in frame.winfo_children():
                if isinstance(rb, tk.Radiobutton):
                    selected = rb.cget("value") == self.period_var.get()
                    rb.config(bg=accent if selected else "#252836",
                              fg="white" if selected else text)
        self.period_var.trace_add("write", lambda *_: update_styles())
        update_styles()

    def _style_radiobuttons_var(self, frame, var, accent, border, text):
        """Apply toggle-button styling to radio buttons for any StringVar."""
        def update_styles(*_):
            for rb in frame.winfo_children():
                if isinstance(rb, tk.Radiobutton):
                    selected = rb.cget("value") == var.get()
                    rb.config(bg=accent if selected else "#252836",
                              fg="white" if selected else text)
        var.trace_add("write", update_styles)
        update_styles()

    def _browse_folder(self):
        folder = filedialog.askdirectory(title="Select output folder")
        if folder:
            self.folder_var.set(folder)

    def _log(self, message: str, tag: str = ""):
        self.log_text.config(state="normal")
        self.log_text.insert("end", message + "\n", tag)
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _start_scrape(self):
        raw = self.ticker_var.get().strip()
        if not raw:
            messagebox.showwarning("No Tickers", "Please enter at least one ticker.")
            return
        tickers = [t.strip().upper() for t in raw.replace(" ", ",").split(",") if t.strip()]
        period_choice = self.period_var.get()
        periods = {"annual": ["annual"], "quarterly": ["quarterly"], "both": ["annual", "quarterly"]}[period_choice]
        fmt = self.format_var.get()
        mode = self.mode_var.get()
        statements = [k for k, v in self.stmt_vars.items() if v.get()]
        if not statements:
            messagebox.showwarning("No Statements", "Please select at least one statement.")
            return
        output_dir = self.folder_var.get().strip() or "stockscraper_output"
        timestamp = datetime.now().strftime("%Y-%m-%d_%I-%M-%S%p").upper()

        self.run_btn.config(state="disabled", text="Running…")
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")
        self.progress["value"] = 0
        self._set_progress_color("blue")

        threading.Thread(target=self._run_scrape,
                         args=(tickers, periods, output_dir, fmt, statements, mode, timestamp),
                         daemon=True).start()

    def _run_scrape(self, tickers, periods, output_dir, fmt, statements, mode, timestamp):
        total = 0
        for t in tickers:
            for s in statements:
                for p in periods:
                    if not (s == "overview" and p == "quarterly"):
                        total += 1
        done = 0
        any_failed = False

        mode_label = "Combined Excel" if mode == "combined" else f"Separate {fmt.upper()}"

        self._log(f"Starting — {len(tickers)} ticker(s), {len(periods)} period(s), {mode_label}", "info")
        self._log(f"Output → {os.path.abspath(output_dir)}\n", "muted")

        # One shared folder for this entire run, containing a subfolder per ticker
        run_dir = os.path.join(output_dir, timestamp)

        for ticker in tickers:
            ticker_dir = os.path.join(run_dir, ticker.upper())
            self._log(f"── {ticker} ──", "info")

            sheets = {}
            ticker_failed = False
            ticker_dir_created = False

            for statement in statements:
                if ticker_failed:
                    break
                for period in periods:
                    if statement == "overview" and period == "quarterly":
                        continue
                    result = scrape_table(ticker, statement, period)
                    label = statement if statement == "overview" else f"{statement} ({period})"

                    if isinstance(result, ScrapeError):
                        self._log(f"  ✗  {label}  — {result.reason}", "fail")
                        self._log(f"  ⚠  Stopping {ticker} — cleaning up", "fail")
                        if ticker_dir_created and os.path.exists(ticker_dir):
                            import shutil
                            shutil.rmtree(ticker_dir)
                            self._log(f"  🗑  Removed folder for {ticker}", "fail")
                        ticker_failed = True
                        any_failed = True
                        break
                    elif result is not None:
                        if not ticker_dir_created:
                            os.makedirs(ticker_dir, exist_ok=True)
                            ticker_dir_created = True
                        if mode == "combined":
                            sheet_map = {
                                "overview":            "Overview",
                                "income-statement":    "Income Stmt",
                                "balance-sheet":       "Balance Sheet",
                                "cash-flow-statement": "Cash Flow",
                                "ratios":              "Ratios",
                            }
                            sheet_name = sheet_map.get(statement, statement)
                            if statement != "overview":
                                sheet_name += f" ({period.capitalize()})"
                            sheets[sheet_name] = result
                            self._log(f"  ✓  {label}  →  sheet: {sheet_name}", "ok")
                        else:
                            path = save_file(result, ticker_dir, ticker, statement, period, fmt)
                            self._log(f"  ✓  {label}  →  {os.path.basename(path)}  ({len(result)} rows)", "ok")
                    else:
                        self._log(f"  ✗  {label}  — no data found", "fail")
                    done += 1
                    self.progress["value"] = (done / total) * 100
                    time.sleep(DELAY_SECONDS)

            if not ticker_failed and mode == "combined" and sheets:
                path = save_combined(sheets, ticker_dir, ticker)
                self._log(f"  📄  Saved → {os.path.basename(path)}  ({len(sheets)} sheets)", "ok")

        self._log(f"\nDone. Files saved to:\n{os.path.abspath(output_dir)}", "info")
        self.run_btn.config(state="normal", text="▶  Run Scraper")
        self.progress["value"] = 100
        self._set_progress_color("red" if any_failed else "green")


if __name__ == "__main__":
    app = App()
    app.mainloop()
